#!/usr/bin/env python3
"""
Script de migração: Pedidos da planilha Clientes (1).xlsx → tabela pedidos (SQLite)

Requer: clientes já migrados (para mapear ClienteID → cliente_id)
Uso:
  cd AtelieIlmaGuerra
  python scripts/migrate_pedidos_excel.py [--dry-run]
  python scripts/migrate_pedidos_excel.py --update-margem   # Atualiza margem_real em pedidos existentes
"""
import sys
from pathlib import Path

backend_path = Path(__file__).resolve().parent.parent / "app_dev" / "backend"
sys.path.insert(0, str(backend_path))

import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from app.core.config import settings
from app.domains.pedidos.models import Pedido, TipoPedido
from app.domains.clientes.models import Cliente
from app.core.database import Base


def _norm(val):
    if val is None:
        return None
    if isinstance(val, str) and val.strip().upper() in ("NA", "N/A", ""):
        return None
    return val


def _norm_float(val):
    v = _norm(val)
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _norm_date(val):
    from datetime import datetime as dt, date as d
    v = _norm(val)
    if v is None:
        return None
    if hasattr(v, "date") and callable(getattr(v, "date")):
        return v.date()
    if isinstance(v, d):
        return v
    if isinstance(v, str):
        try:
            parsed = dt.fromisoformat(v.replace("Z", "+00:00"))
            return parsed.date()
        except (ValueError, TypeError):
            return None
    return None


def _norm_status(val):
    v = _norm(val)
    if not v:
        return "Encomenda"
    s = str(v).strip()
    valid = ("Orçamento", "Encomenda", "Cortado", "Provado", "Pronto", "Entregue", "Canelado", "Backlog")
    if s in valid:
        return "Encomenda" if s == "Backlog" else s
    return "Encomenda"


def run_migration(excel_path: Path, dry_run: bool = False):
    engine = create_engine(settings.DATABASE_URL)
    SessionLocal = sessionmaker(bind=engine)
    Base.metadata.create_all(bind=engine)
    add_fotos_columns_if_needed()
    add_param_columns_if_needed()

    db = SessionLocal()

    # Mapa appsheet_id -> cliente_id
    clientes_map = {c.appsheet_id: c.id for c in db.query(Cliente).filter(Cliente.appsheet_id.isnot(None)).all()}
    # Mapa nome tipo -> tipo_pedido_id
    tipos_map = {t.nome.upper(): t.id for t in db.query(TipoPedido).all()}

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Pedidos"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        print("Planilha Pedidos vazia.")
        db.close()
        return

    header = [str(c).strip() if c else "" for c in rows[0]]
    col_map = {h: i for i, h in enumerate(header)}

    def get(row, name, default=None):
        idx = col_map.get(name)
        if idx is None or idx >= len(row):
            return default
        return row[idx]

    created = 0
    skipped = 0
    orphan = 0

    for i, row in enumerate(rows[1:], start=2):
        appsheet_id = _norm(get(row, "PedidoID"))
        cliente_appsheet = _norm(get(row, "ClienteID"))
        tipo_nome = _norm(get(row, "TipoPedido"))
        data_pedido = _norm_date(get(row, "DataPedido"))
        data_entrega = _norm_date(get(row, "DataEntrega"))

        if not data_pedido:
            skipped += 1
            continue

        cliente_id = clientes_map.get(cliente_appsheet) if cliente_appsheet else None
        if not cliente_id:
            orphan += 1
            continue

        tipo_id = tipos_map.get(str(tipo_nome).upper()) if tipo_nome else None

        if appsheet_id:
            existing = db.query(Pedido).filter(Pedido.appsheet_id == appsheet_id).first()
            if existing:
                skipped += 1
                continue

        try:
            valor_pecas = _norm_float(get(row, "ValorPecas")) or _norm_float(get(row, "ValorTotal"))
            margem_raw = _norm_float(get(row, "MargemReal_1")) or _norm_float(get(row, "MargemReal"))
            # MargemReal pode vir como 34 (%) ou 0.34; normalizar para % (34)
            margem_real = None
            if margem_raw is not None:
                margem_real = margem_raw if margem_raw >= 1 or margem_raw == 0 else margem_raw * 100
            flag_med = _norm(get(row, "FlagMedidas"))
            medidas_disp = flag_med and str(flag_med).strip().upper() in ("SIM", "S", "1")
            pedido = Pedido(
                appsheet_id=appsheet_id,
                cliente_id=cliente_id,
                tipo_pedido_id=tipo_id,
                data_pedido=data_pedido,
                data_entrega=data_entrega,
                descricao_produto=str(get(row, "DescricaoProduto") or "").strip() or "",
                status=_norm_status(get(row, "Status (Backlog/Cortado/Pronto/Entregue)")),
                valor_pecas=valor_pecas,
                quantidade_pecas=int(x) if (x := _norm_float(get(row, "Quantidade"))) and x == int(x) else None,
                horas_trabalho=_norm_float(get(row, "HorasTrabalho")),
                custo_materiais=_norm_float(get(row, "custo_materiais")),
                custos_variaveis=_norm_float(get(row, "custos_variaveis")),
                margem_real=margem_real,
                forma_pagamento=_norm(get(row, "FormaPagamento")),
                valor_entrada=_norm_float(get(row, "Entrada")),
                valor_restante=_norm_float(get(row, "ValorRestante")),
                detalhes_pagamento=_norm(get(row, "DetalhePagamento")),
                medidas_disponiveis=medidas_disp,
                medida_ombro=_norm_float(get(row, "MedidaOmbro")),
                medida_busto=_norm_float(get(row, "MedidaBusto")),
                medida_cinto=_norm_float(get(row, "MedidaCinto")),
                medida_quadril=_norm_float(get(row, "MedidaQuadril")),
                medida_comprimento_corpo=_norm_float(get(row, "MedidaComprimentoCorpo")),
                medida_comprimento_vestido=_norm_float(get(row, "MedidaComprimentoVestido")),
                medida_seio_a_seio=_norm_float(get(row, "MedidaSeioaSeio")),
                medida_raio_busto=_norm_float(get(row, "MedidaRaioDeBusto")),
                medida_altura_busto=_norm_float(get(row, "MedidaAlturaDeBusto")),
                medida_frente=_norm_float(get(row, "MedidaFrente")),
                medida_costado=_norm_float(get(row, "MedidaCostado")),
                medida_comprimento_calca=_norm_float(get(row, "MedidaComprimentoCalca")),
                medida_comprimento_blusa=_norm_float(get(row, "MedidaComprimentoBlusa")),
                medida_largura_manga=_norm_float(get(row, "MedidaLarguraDaManga")),
                medida_comprimento_manga=_norm_float(get(row, "MedidaComprimentoDaManga")),
                medida_punho=_norm_float(get(row, "MedidaPunho")),
                medida_comprimento_saia=_norm_float(get(row, "MedidaComprimentoDaSaia")),
                medida_comprimento_bermuda=_norm_float(get(row, "MedidaComprimentoDaBermuda")),
            )
            if not dry_run:
                db.add(pedido)
            created += 1
        except Exception as e:
            print(f"  Linha {i}: {e}")

    if not dry_run:
        db.commit()
    db.close()

    print(f"Migração Pedidos: {created} criados, {skipped} ignorados, {orphan} órfãos (cliente não encontrado).")
    if dry_run:
        print("(dry-run: nenhum dado gravado)")


def update_margem_from_excel(excel_path: Path, dry_run: bool = False):
    """Atualiza margem_real em pedidos existentes a partir do Excel."""
    engine = create_engine(settings.DATABASE_URL)
    SessionLocal = sessionmaker(bind=engine)
    db = SessionLocal()

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Pedidos"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = [str(c).strip() if c else "" for c in rows[0]]
    col_map = {h: i for i, h in enumerate(header)}

    def get(row, name, default=None):
        idx = col_map.get(name)
        if idx is None or idx >= len(row):
            return default
        return row[idx]

    updated = 0
    for row in rows[1:]:
        appsheet_id = _norm(get(row, "PedidoID"))
        if not appsheet_id:
            continue
        margem_raw = _norm_float(get(row, "MargemReal_1")) or _norm_float(get(row, "MargemReal"))
        margem_real = None
        if margem_raw is not None:
            margem_real = margem_raw if margem_raw >= 1 or margem_raw == 0 else margem_raw * 100
        pedido = db.query(Pedido).filter(Pedido.appsheet_id == appsheet_id).first()
        if pedido and pedido.margem_real != margem_real:
            if not dry_run:
                pedido.margem_real = margem_real
            updated += 1

    if not dry_run:
        db.commit()
    db.close()
    print(f"Margem atualizada em {updated} pedidos.")
    if dry_run:
        print("(dry-run: nenhum dado gravado)")


def add_param_columns_if_needed():
    """Adiciona colunas de snapshot de parâmetros usados no cálculo do pedido."""
    engine = create_engine(settings.DATABASE_URL)
    with engine.connect() as conn:
        info = conn.execute(text("PRAGMA table_info(pedidos)"))
        cols = {row[1] for row in info}
    new_cols = [
        "param_preco_hora", "param_impostos", "param_cartao_credito",
        "param_total_horas_mes", "param_margem_target",
    ]
    added = 0
    with engine.connect() as conn:
        for col in new_cols:
            if col not in cols:
                conn.execute(text(f"ALTER TABLE pedidos ADD COLUMN {col} REAL"))
                conn.commit()
                added += 1
    if added:
        print(f"Colunas de parâmetros adicionadas: {added}")
    return added


def add_fotos_columns_if_needed():
    """Adiciona colunas de fotos 2, 3 e comentários se não existirem."""
    engine = create_engine(settings.DATABASE_URL)
    with engine.connect() as conn:
        info = conn.execute(text("PRAGMA table_info(pedidos)"))
        cols = {row[1] for row in info}
    new_cols = [
        "foto_url_2", "foto_url_3",
        "comentario_foto_1", "comentario_foto_2", "comentario_foto_3",
    ]
    added = 0
    with engine.connect() as conn:
        for col in new_cols:
            if col not in cols:
                tipo = "TEXT" if "comentario" in col else "VARCHAR(500)"
                conn.execute(text(f"ALTER TABLE pedidos ADD COLUMN {col} {tipo}"))
                conn.commit()
                added += 1
    if added:
        print(f"Colunas de fotos adicionadas: {added}")
    return added


def add_medidas_columns_if_needed():
    """Adiciona colunas de medidas extras na tabela pedidos se não existirem."""
    engine = create_engine(settings.DATABASE_URL)
    with engine.connect() as conn:
        info = conn.execute(text("PRAGMA table_info(pedidos)"))
        cols = {row[1] for row in info}
    new_cols = [
        "medida_quadril", "medida_comprimento_corpo", "medida_comprimento_vestido",
        "medida_seio_a_seio", "medida_raio_busto", "medida_altura_busto",
        "medida_frente", "medida_costado", "medida_comprimento_calca",
        "medida_comprimento_blusa", "medida_largura_manga", "medida_comprimento_manga",
        "medida_punho", "medida_comprimento_saia", "medida_comprimento_bermuda",
    ]
    added = 0
    with engine.connect() as conn:
        for col in new_cols:
            if col not in cols:
                conn.execute(text(f"ALTER TABLE pedidos ADD COLUMN {col} REAL"))
                conn.commit()
                added += 1
    if added:
        print(f"Colunas de medidas adicionadas: {added}")
    return added


def update_medidas_from_excel(excel_path: Path, dry_run: bool = False):
    """Atualiza medidas em pedidos existentes a partir do Excel."""
    engine = create_engine(settings.DATABASE_URL)
    SessionLocal = sessionmaker(bind=engine)
    db = SessionLocal()

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Pedidos"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = [str(c).strip() if c else "" for c in rows[0]]
    col_map = {h: i for i, h in enumerate(header)}

    def get(row, name, default=None):
        idx = col_map.get(name)
        if idx is None or idx >= len(row):
            return default
        return row[idx]

    def _med(name):
        return _norm_float(get(row, name))

    updated = 0
    for row in rows[1:]:
        appsheet_id = _norm(get(row, "PedidoID"))
        if not appsheet_id:
            continue
        flag_med = _norm(get(row, "FlagMedidas"))
        medidas_disp = flag_med and str(flag_med).strip().upper() in ("SIM", "S", "1")
        meds = {
            "medidas_disponiveis": medidas_disp,
            "medida_ombro": _med("MedidaOmbro"),
            "medida_busto": _med("MedidaBusto"),
            "medida_cinto": _med("MedidaCinto"),
            "medida_quadril": _med("MedidaQuadril"),
            "medida_comprimento_corpo": _med("MedidaComprimentoCorpo"),
            "medida_comprimento_vestido": _med("MedidaComprimentoVestido"),
            "medida_seio_a_seio": _med("MedidaSeioaSeio"),
            "medida_raio_busto": _med("MedidaRaioDeBusto"),
            "medida_altura_busto": _med("MedidaAlturaDeBusto"),
            "medida_frente": _med("MedidaFrente"),
            "medida_costado": _med("MedidaCostado"),
            "medida_comprimento_calca": _med("MedidaComprimentoCalca"),
            "medida_comprimento_blusa": _med("MedidaComprimentoBlusa"),
            "medida_largura_manga": _med("MedidaLarguraDaManga"),
            "medida_comprimento_manga": _med("MedidaComprimentoDaManga"),
            "medida_punho": _med("MedidaPunho"),
            "medida_comprimento_saia": _med("MedidaComprimentoDaSaia"),
            "medida_comprimento_bermuda": _med("MedidaComprimentoDaBermuda"),
        }
        pedido = db.query(Pedido).filter(Pedido.appsheet_id == appsheet_id).first()
        if not pedido:
            continue
        changed = any(getattr(pedido, k, None) != v for k, v in meds.items())
        if changed:
            if not dry_run:
                for k, v in meds.items():
                    setattr(pedido, k, v)
            updated += 1

    if not dry_run:
        db.commit()
    db.close()
    print(f"Medidas atualizadas em {updated} pedidos.")
    if dry_run:
        print("(dry-run: nenhum dado gravado)")


if __name__ == "__main__":
    project_root = Path(__file__).resolve().parent.parent
    excel_path = project_root / "Clientes (1).xlsx"
    if not excel_path.exists():
        print(f"Arquivo não encontrado: {excel_path}")
        sys.exit(1)
    do_margem = "--update-margem" in sys.argv
    do_medidas = "--update-medidas" in sys.argv
    dry_run = "--dry-run" in sys.argv
    if do_margem:
        update_margem_from_excel(excel_path, dry_run=dry_run)
    if do_medidas:
        add_medidas_columns_if_needed()
        update_medidas_from_excel(excel_path, dry_run=dry_run)
    if not do_margem and not do_medidas:
        run_migration(excel_path, dry_run=dry_run)
