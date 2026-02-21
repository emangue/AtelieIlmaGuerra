#!/usr/bin/env python3
"""
Script de migração: ParametrosOrcamento1 da planilha Clientes (1).xlsx → tabela despesas_detalhadas

A planilha ParametrosOrcamento1 contém as despesas detalhadas (GrupoDespesa, DetalheDespesa, Valor)
que alimentam o Total Despesas nos parâmetros.

Uso:
  cd AtelieIlmaGuerra
  python scripts/migrate_despesas_excel.py [--dry-run]
"""
import sys
from pathlib import Path

backend_path = Path(__file__).resolve().parent.parent / "app_dev" / "backend"
sys.path.insert(0, str(backend_path))

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from sqlalchemy import text

from app.core.config import settings
from app.core.database import Base
from app.domains.despesas.models import DespesaDetalhada


def add_appsheet_id_column_if_needed():
    """Adiciona coluna appsheet_id na tabela despesas_detalhadas se não existir."""
    engine = create_engine(settings.DATABASE_URL)
    with engine.connect() as conn:
        info = conn.execute(text("PRAGMA table_info(despesas_detalhadas)"))
        cols = {row[1] for row in info}
    if "appsheet_id" not in cols:
        with engine.connect() as conn:
            conn.execute(text("ALTER TABLE despesas_detalhadas ADD COLUMN appsheet_id VARCHAR(50)"))
            conn.commit()
        print("Coluna appsheet_id adicionada em despesas_detalhadas.")


def _norm(val):
    if val is None:
        return None
    if isinstance(val, str) and val.strip().upper() in ("NA", "N/A", ""):
        return None
    return val


def _norm_float(val):
    v = _norm(val)
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def run_migration(excel_path: Path, dry_run: bool = False):
    engine = create_engine(settings.DATABASE_URL)
    SessionLocal = sessionmaker(bind=engine)
    Base.metadata.create_all(bind=engine)
    add_appsheet_id_column_if_needed()

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "ParametrosOrcamento1" not in wb.sheetnames:
        print("Planilha ParametrosOrcamento1 não encontrada.")
        wb.close()
        return

    ws = wb["ParametrosOrcamento1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        print("Planilha ParametrosOrcamento1 vazia.")
        return

    header = [str(c).strip() if c else "" for c in rows[0]]
    col_map = {h: i for i, h in enumerate(header)}

    def get(row, name, default=None):
        idx = col_map.get(name)
        if idx is None or idx >= len(row):
            return default
        return row[idx]

    db = SessionLocal()
    created = 0
    updated = 0
    skipped = 0

    for row in rows[1:]:
        appsheet_id = _norm(get(row, "DespesaID"))
        detalhe = _norm(get(row, "DetalheDespesa"))
        valor = _norm_float(get(row, "Valor"))
        grupo = _norm(get(row, "GrupoDespesa"))

        if not detalhe:
            skipped += 1
            continue
        if valor is None or valor < 0:
            valor = 0

        existing = None
        if appsheet_id:
            existing = db.query(DespesaDetalhada).filter(DespesaDetalhada.appsheet_id == appsheet_id).first()
        if not existing and detalhe:
            existing = (
                db.query(DespesaDetalhada)
                .filter(
                    DespesaDetalhada.detalhe == str(detalhe).strip(),
                    DespesaDetalhada.grupo == (str(grupo).strip() if grupo else None),
                )
                .first()
            )

        if existing:
            if (existing.valor != valor or (appsheet_id and not existing.appsheet_id)) and not dry_run:
                existing.valor = valor
                if appsheet_id:
                    existing.appsheet_id = appsheet_id
                db.commit()
                updated += 1
            else:
                skipped += 1
            continue

        if not dry_run:
            d = DespesaDetalhada(
                appsheet_id=appsheet_id,
                detalhe=str(detalhe).strip(),
                valor=valor,
                grupo=str(grupo).strip() if grupo else None,
            )
            db.add(d)
            created += 1

    if not dry_run:
        db.commit()
    db.close()

    print(f"Migração ParametrosOrcamento1 → despesas_detalhadas: {created} criados, {updated} atualizados, {skipped} ignorados.")
    if dry_run:
        print("(dry-run: nenhum dado gravado)")


if __name__ == "__main__":
    project_root = Path(__file__).resolve().parent.parent
    excel_path = project_root / "Clientes (1).xlsx"
    if not excel_path.exists():
        print(f"Arquivo não encontrado: {excel_path}")
        sys.exit(1)
    dry_run = "--dry-run" in sys.argv
    run_migration(excel_path, dry_run=dry_run)
